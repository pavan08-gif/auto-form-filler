from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
wb = load_workbook(filename = 'peoples_data.xlsx')
ws = wb.active
driver = webdriver.Chrome()
wait = WebDriverWait(driver,10)
for row in ws.iter_rows(max_col = 3,min_row = 2,max_row = 11,values_only = True):
    driver.get("https://docs.google.com/forms/d/e/1FAIpQLSegf891KAK7q_F2RIFrc8Grp03ygOFHD3vWjoNiXnUQbDGgRQ/viewform")
    #element_list = wait.until(lambda d: d.find_elements(By.CSS_SELECTOR,"input.whsOnd.zHQkBf")
    #if (len(d.find_elements(By.CSS_SELECTOR,"input.whsOnd.zHQkBf"))==3)
    #else False)
    #element_name = element_list[0]
    #element_age = element_list[1]
    #element_email = element_list[2]
    element_name = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[aria-labelledby = 'i1 i4']")))
    #driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element_name)
    #time.sleep(3)
    element_name.click()
    element_name.send_keys(row[0])
    element_age = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"input[aria-labelledby = 'i6 i9']")))
    element_age.click()
    element_age.send_keys(str(row[1]))
    element_email = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"input[aria-labelledby = 'i11 i14']")))
    element_email.click()
    element_email.send_keys(row[2])
    element_submit = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,'div[aria-label = "Submit"]')))
    #driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element_submit)
    element_submit.click()
#testing if git works
driver.quit()

 